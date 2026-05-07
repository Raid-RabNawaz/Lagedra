from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def add_table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def format_sheet(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate release gantt Excel workbook")
    parser.add_argument("--output", default="RELEASE_GANTT_2026-04.xlsx", help="Output Excel filename")
    args = parser.parse_args()

    output = Path(__file__).resolve().parents[1] / args.output

    wb = Workbook()
    ws_items = wb.active
    ws_items.title = "Item_Level"

    ws_items.append(
        [
            "Release",
            "Track",
            "Item",
            "Status",
            "Priority",
            "Dependency Type",
            "Start Date",
            "Due Date (Wednesday)",
            "Owner",
            "Notes",
        ]
    )

    item_rows = [
        ("R1", "Listings", "Book Now click interception fix", "In Progress", "P0", "Internal", date(2026, 4, 20), date(2026, 4, 22), "", ""),
        ("R1", "Reservations", "Reservations route alignment", "In Progress", "P1", "Internal", date(2026, 4, 20), date(2026, 4, 22), "", ""),
        ("R1", "Listings", "Similar listings 400 + fallback", "In Progress", "P1", "Internal", date(2026, 4, 20), date(2026, 4, 22), "", ""),
        ("R1", "Search", "Search keyword filtering patch", "In Progress", "P1", "Internal", date(2026, 4, 20), date(2026, 4, 22), "", ""),
        ("R1", "Verification", "Identity/risk status 404 fix", "In Progress", "P0", "Internal", date(2026, 4, 21), date(2026, 4, 22), "", ""),
        ("R1", "Verification", "KYC start 451 fix", "In Progress", "P0", "Third-Party API", date(2026, 4, 21), date(2026, 4, 22), "", ""),
        ("R1", "Billing", "Checkout 400 + UI error handling", "In Progress", "P0", "Third-Party API", date(2026, 4, 21), date(2026, 4, 22), "", ""),
        ("R1", "Release Ops", "Smoke test + release notes", "Planned", "P1", "Internal", date(2026, 4, 21), date(2026, 4, 22), "", ""),
        ("R2", "Frontend UX", "Session/header flicker reduction", "Planned", "P1", "Internal", date(2026, 4, 23), date(2026, 4, 29), "", ""),
        ("R2", "Frontend UX", "Empty/error state UX hardening", "Planned", "P1", "Internal", date(2026, 4, 23), date(2026, 4, 29), "", ""),
        ("R2", "Admin", "Admin panel stabilization", "Planned", "P1", "Internal", date(2026, 4, 23), date(2026, 4, 29), "", ""),
        ("R2", "Applications", "Application flow improvements", "Planned", "P1", "Internal", date(2026, 4, 23), date(2026, 4, 29), "", ""),
        ("R2", "Ops", "Deployment docs finalization", "Planned", "P2", "Internal", date(2026, 4, 24), date(2026, 4, 29), "", ""),
        ("R2", "Realtime", "Websocket retry + health indicator", "Planned", "P2", "Third-Party API", date(2026, 4, 25), date(2026, 4, 29), "", ""),
        ("R2", "Billing", "Checkout refinements", "Planned", "P1", "Third-Party API", date(2026, 4, 25), date(2026, 4, 29), "", ""),
        ("R2", "Release Ops", "Regression + release gate checks", "Planned", "P1", "Internal", date(2026, 4, 28), date(2026, 4, 29), "", ""),
        ("R3", "Stability", "Post-R2 bug burn-down", "Planned", "P1", "Internal", date(2026, 4, 30), date(2026, 5, 6), "", ""),
        ("R3", "Observability", "Observability and alert tuning", "Planned", "P2", "Internal", date(2026, 4, 30), date(2026, 5, 6), "", ""),
        ("R3", "Performance", "Performance and reliability pass", "Planned", "P2", "Internal", date(2026, 4, 30), date(2026, 5, 6), "", ""),
        ("R3", "Release Ops", "R3 release readiness review", "Planned", "P1", "Internal", date(2026, 5, 5), date(2026, 5, 6), "", ""),
    ]

    for row in item_rows:
        ws_items.append(row)

    for row in ws_items.iter_rows(min_row=2, min_col=7, max_col=8):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"

    add_table(ws_items, f"A1:J{ws_items.max_row}", "ItemLevelTable")
    format_sheet(
        ws_items,
        {"A": 10, "B": 16, "C": 42, "D": 14, "E": 10, "F": 18, "G": 14, "H": 20, "I": 16, "J": 32},
    )

    ws_module = wb.create_sheet("Module_Progress")
    ws_module.append(
        [
            "Module",
            "Release",
            "Task",
            "Status",
            "Start Date",
            "Due Date (Wednesday)",
            "Progress %",
            "Notes",
        ]
    )

    module_rows = [
        ("Activation and Billing", "R1", "Internal checkout UI/state fixes", "In Progress", date(2026, 4, 20), date(2026, 4, 22), 60, ""),
        ("Activation and Billing", "R1", "Checkout gateway integration (Third-Party API)", "Planned", date(2026, 4, 21), date(2026, 4, 22), 20, ""),
        ("Activation and Billing", "R2", "Checkout refinements (Third-Party API)", "Planned", date(2026, 4, 25), date(2026, 4, 29), 10, ""),
        ("Verification and Risk", "R1", "Identity/risk endpoint fixes", "In Progress", date(2026, 4, 20), date(2026, 4, 22), 55, ""),
        ("Verification and Risk", "R1", "KYC vendor flow fix (Third-Party API)", "Planned", date(2026, 4, 21), date(2026, 4, 22), 20, ""),
        ("Verification and Risk", "R2", "Verification UX hardening", "Planned", date(2026, 4, 23), date(2026, 4, 29), 10, ""),
        ("Listings and Booking", "R1", "Book now click + route fixes", "In Progress", date(2026, 4, 20), date(2026, 4, 22), 50, ""),
        ("Listings and Booking", "R1", "Similar/search API behavior fixes", "In Progress", date(2026, 4, 20), date(2026, 4, 22), 45, ""),
        ("Listings and Booking", "R2", "Listing UX polish", "Planned", date(2026, 4, 23), date(2026, 4, 29), 5, ""),
        ("Applications", "R2", "Apply dialog/detail fixes", "Planned", date(2026, 4, 23), date(2026, 4, 29), 10, ""),
        ("Applications", "R2", "Empty/error state improvements", "Planned", date(2026, 4, 23), date(2026, 4, 29), 10, ""),
        ("Admin and Operations", "R2", "Admin pages stabilization", "Planned", date(2026, 4, 23), date(2026, 4, 29), 10, ""),
        ("Admin and Operations", "R2", "Consent + deployment documentation", "Planned", date(2026, 4, 24), date(2026, 4, 29), 20, ""),
        ("Admin and Operations", "R2", "Release gates and runbook checks", "Planned", date(2026, 4, 28), date(2026, 4, 29), 0, ""),
    ]
    for row in module_rows:
        ws_module.append(row)

    for row in ws_module.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"
    for cell in ws_module["G"][1:]:
        cell.number_format = "0%"
        cell.value = (cell.value or 0) / 100

    add_table(ws_module, f"A1:H{ws_module.max_row}", "ModuleProgressTable")
    format_sheet(
        ws_module,
        {"A": 24, "B": 10, "C": 40, "D": 14, "E": 14, "F": 20, "G": 12, "H": 30},
    )

    ws_targets = wb.create_sheet("Release_Targets")
    ws_targets.append(["Release", "Release Date (Wednesday)", "Target Completion", "Target Scope", "Current Status"])
    target_rows = [
        ("R1", date(2026, 4, 22), "100%", "All blocker fixes completed", "On Track"),
        ("R2", date(2026, 4, 29), "100%", "UX stabilization + admin + docs", "At Risk"),
        ("R3", date(2026, 5, 6), "100%", "Hardening (reliability/observability/perf)", "Planned"),
    ]
    for row in target_rows:
        ws_targets.append(row)

    for cell in ws_targets["B"][1:]:
        cell.number_format = "yyyy-mm-dd"

    add_table(ws_targets, f"A1:E{ws_targets.max_row}", "ReleaseTargetsTable")
    format_sheet(ws_targets, {"A": 10, "B": 26, "C": 20, "D": 42, "E": 16})

    ws_notes = wb.create_sheet("How_To_Update")
    ws_notes.append(["How to Keep This Updated"])
    notes = [
        "1) Update Status, Owner, and Notes in Item_Level each day.",
        "2) Set Due Date to the correct Wednesday release date.",
        "3) Update Progress % in Module_Progress once per standup.",
        "4) If an item slips, change Release and Due Date to next Wednesday.",
        "5) Add new rows directly in the tables; Excel keeps formatting.",
    ]
    for note in notes:
        ws_notes.append([note])
    ws_notes.column_dimensions["A"].width = 90
    ws_notes["A1"].font = Font(bold=True)

    wb.save(output)
    print(f"Created: {output}")


if __name__ == "__main__":
    main()
